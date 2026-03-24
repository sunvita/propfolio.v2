"""
Propfolio AU — Excel Generator Service.

Generates professionally formatted .xlsx workbooks matching the exact
Australian property P&L template:
  - Per-property IP# sheets (13-column FY blocks, reverse-chrono months)
  - Summary sheet with cross-sheet formulas
  - KPI section per property
  - Full styling, number formats, formulas, print setup
"""

import re
from pathlib import Path
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

from backend.config import (
    OUTPUT_DIR,
    INCOME_ROWS, INCOME_TOTAL_ROW,
    OPEX_ROWS, OPEX_TOTAL_ROW,
    NOI_ROW, NOI_MARGIN_ROW,
    UTILITY_ROWS, UTILITY_TOTAL_ROW,
    FINANCING_ROWS, FINANCING_TOTAL_ROW,
    CAPITAL_ROWS, CAPITAL_TOTAL_ROW,
    NET_PROFIT_ROW,
    CASHFLOW_ROWS, NET_CASHFLOW_ROW, PRINCIPAL_ROW,
    KPI_HEADER_ROW, KPI_START_ROW,
    SECTION_HEADERS, CATEGORY_ROW_MAP,
    COLORS, NUM_FMT_CURRENCY, NUM_FMT_PERCENT, NUM_FMT_RATIO,
    FY_BLOCK_WIDTH, MONTHS_IN_FY,
)
from backend.services.fy_utils import (
    get_fy, get_fy_year_range, get_fy_months,
    get_fy_list_from_transactions, get_cy_list_from_transactions,
)
from backend.services.ledger import (
    load_ledger, load_portfolio, aggregate_by_category_month,
)


# ── Reusable Styles ──────────────────────────────────────────────────

def _rgb(hex_str: str) -> str:
    """Convert 6-char hex to 8-char ARGB (openpyxl uses 'FF' prefix)."""
    return f"FF{hex_str}"


def _fill(color_key: str) -> PatternFill:
    return PatternFill(start_color=_rgb(COLORS[color_key]),
                       end_color=_rgb(COLORS[color_key]),
                       fill_type="solid")


def _font(color_key: str, size: int = 9, bold: bool = False,
          italic: bool = False, name: str = "Arial") -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic,
                color=_rgb(COLORS[color_key]))


THIN_SIDE = Side(style="thin")
MEDIUM_SIDE = Side(style="medium")
DOUBLE_SIDE = Side(style="double")
THICK_SIDE = Side(style="thick")

BORDER_TOP_THIN = Border(top=THIN_SIDE)
BORDER_TOP_BOTTOM_MEDIUM = Border(top=MEDIUM_SIDE, bottom=MEDIUM_SIDE)
BORDER_TOP_THIN_BOTTOM_DOUBLE = Border(top=THIN_SIDE, bottom=DOUBLE_SIDE)
BORDER_TOP_BOTTOM_THICK = Border(top=THICK_SIDE, bottom=THICK_SIDE)


# ── Column helpers ────────────────────────────────────────────────────

def _fy_total_col(fy_index: int) -> int:
    """Return the 1-based column number for the FY Total column.
    fy_index 0 = most recent FY (column B=2).
    """
    return 2 + fy_index * FY_BLOCK_WIDTH


def _month_col(fy_index: int, month_offset: int) -> int:
    """Return 1-based column for a month within an FY block.
    month_offset 0 = first month after FY Total (Jun, the most recent).
    """
    return 3 + fy_index * FY_BLOCK_WIDTH + month_offset


def _month_key_from_fy(fy_label: str, month_offset: int) -> str:
    """Get 'YYYY-MM' key for a month_offset in an FY block.
    Months are in reverse-chrono order: Jun, May, Apr, ..., Jul.
    """
    start_year, end_year = get_fy_year_range(fy_label)
    # MONTHS_IN_FY = ["Jun","May","Apr","Mar","Feb","Jan","Dec","Nov","Oct","Sep","Aug","Jul"]
    month_name = MONTHS_IN_FY[month_offset]
    month_num = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
    }[month_name]
    year = end_year if month_num <= 6 else start_year
    return f"{year}-{month_num:02d}"


# ── Main API ──────────────────────────────────────────────────────────

def generate_workbook(
    property_ids: Optional[list[str]] = None,
    output_filename: Optional[str] = None,
) -> str:
    """Generate the full Excel workbook for the portfolio.

    Args:
        property_ids: List of property IDs to include. None = all.
        output_filename: Custom filename. Auto-generated if None.

    Returns:
        Path to the generated .xlsx file.
    """
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    portfolio = load_portfolio()
    props = portfolio.properties

    if property_ids:
        props = [p for p in props if p.id in property_ids]

    if not props:
        raise ValueError("No properties found to generate report.")

    # Determine the union of all FYs across all properties
    all_fy_labels = set()
    prop_aggregates = {}  # prop_id → {(category, "YYYY-MM"): amount}

    for prop in props:
        agg = aggregate_by_category_month(prop.id)
        prop_aggregates[prop.id] = agg
        ledger = load_ledger(prop.id)
        tx_dicts = [t.model_dump(mode="json") for t in ledger.transactions]
        fy_list = get_fy_list_from_transactions(tx_dicts)
        all_fy_labels.update(fy_list)

    if not all_fy_labels:
        # No transactions at all — create at least current FY
        current_fy = get_fy(datetime.now())
        all_fy_labels.add(current_fy)

    fy_labels = sorted(all_fy_labels, reverse=True)  # Most recent first
    num_fys = len(fy_labels)

    # Build CY list from FY boundaries
    cy_years = set()
    for fy in fy_labels:
        s, e = get_fy_year_range(fy)
        cy_years.add(str(s))
        cy_years.add(str(e))
    cy_labels = sorted(cy_years, reverse=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Build per-property sheets
    for prop in props:
        sheet_name = f"IP#{prop.id.replace('IP', '')}"
        ws = wb.create_sheet(title=sheet_name)
        _build_property_sheet(
            ws, prop, fy_labels, cy_labels,
            prop_aggregates.get(prop.id, {}),
        )

    # Build Summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    _build_summary_sheet(ws_summary, props, fy_labels, cy_labels)

    # Save
    if not output_filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Propfolio_PnL_{ts}.xlsx"
    if not output_filename.endswith(".xlsx"):
        output_filename += ".xlsx"

    out_path = Path(OUTPUT_DIR) / output_filename
    wb.save(str(out_path))
    return str(out_path)


# ══════════════════════════════════════════════════════════════════════
#  PROPERTY SHEET BUILDER
# ══════════════════════════════════════════════════════════════════════

def _build_property_sheet(
    ws, prop, fy_labels: list[str], cy_labels: list[str],
    agg: dict,
):
    """Build a single IP# sheet with full P&L, styling, and formulas."""

    num_fys = len(fy_labels)
    last_data_col = 1 + num_fys * FY_BLOCK_WIDTH  # Last FY month column
    # CY columns come after
    cy_start_col = last_data_col + 1
    total_cols = cy_start_col + len(cy_labels) - 1

    # ── Row 1: Title bar ──────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=prop.display_name)
    title_cell.fill = _fill("title_bg")
    title_cell.font = Font(name="Arial", size=12, bold=True, color=_rgb(COLORS["title_font"]))
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Row 2: Legend ─────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    legend_text = (
        "Green = Income  |  Red = Expense  |  Blue = Computed  |  "
        "Purple = Cash Flow  |  Yellow = Active FY  |  Grey = Past FY"
    )
    legend_cell = ws.cell(row=2, column=1, value=legend_text)
    legend_cell.fill = _fill("legend_bg")
    legend_cell.font = Font(name="Arial", size=8, color=_rgb(COLORS["legend_font"]))
    legend_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    # ── Row 3: Blank ──────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Column headers ─────────────────────────────────────────
    ws.row_dimensions[4].height = 34

    # Col A header
    hdr_a = ws.cell(row=4, column=1, value="Category")
    hdr_a.fill = _fill("header_bg")
    hdr_a.font = Font(name="Arial", size=9, bold=True, color=_rgb(COLORS["header_font"]))
    hdr_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_a.border = Border(bottom=THIN_SIDE)

    # FY block headers
    for fi, fy in enumerate(fy_labels):
        is_active = (fi == 0)  # Most recent FY is "active"

        # FY Total column header
        ft_col = _fy_total_col(fi)
        ft_cell = ws.cell(row=4, column=ft_col, value=f"{fy}\nTotal")
        ft_cell.fill = _fill("fy_total_bg")
        ft_cell.font = Font(name="Arial", size=8, bold=True, color="FF000000")
        ft_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ft_cell.border = Border(bottom=THIN_SIDE)

        # Month columns
        start_year, end_year = get_fy_year_range(fy)
        for mi in range(12):
            mc = _month_col(fi, mi)
            month_name = MONTHS_IN_FY[mi]
            month_num = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                         "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}[month_name]
            yr = end_year if month_num <= 6 else start_year
            label = f"{month_name}\n{yr}"

            mc_cell = ws.cell(row=4, column=mc, value=label)
            if is_active:
                mc_cell.fill = _fill("active_fy_bg")
                mc_cell.font = Font(name="Arial", size=8, bold=True, color="FF000000")
            else:
                mc_cell.fill = _fill("inactive_bg")
                mc_cell.font = Font(name="Arial", size=8, bold=True,
                                    color=_rgb(COLORS["inactive_font"]))
            mc_cell.alignment = Alignment(horizontal="center", vertical="center",
                                          wrap_text=True)
            mc_cell.border = Border(bottom=THIN_SIDE)

    # CY column headers
    for ci, cy in enumerate(cy_labels):
        cc = cy_start_col + ci
        cc_cell = ws.cell(row=4, column=cc, value=f"CY\n{cy}")
        cc_cell.fill = _fill("fy_total_bg")
        cc_cell.font = Font(name="Arial", size=8, bold=True, color="FF000000")
        cc_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc_cell.border = Border(bottom=THIN_SIDE)

    # ── Section headers ───────────────────────────────────────────────
    for row_num, section_name in SECTION_HEADERS.items():
        _write_section_header(ws, row_num, section_name, total_cols)

    # ── Data rows (Income) ────────────────────────────────────────────
    for cat_key, (row, display) in INCOME_ROWS.items():
        _write_data_row(ws, row, display, cat_key, "income",
                        fy_labels, cy_labels, cy_start_col, agg)

    # Income Total (row 9) — SUM formula
    _write_total_row(
        ws, INCOME_TOTAL_ROW, "Total Income",
        [r for _, (r, _) in INCOME_ROWS.items()],
        fy_labels, cy_labels, cy_start_col, "income_total",
    )

    # ── Data rows (OpEx) ──────────────────────────────────────────────
    for cat_key, (row, display) in OPEX_ROWS.items():
        _write_data_row(ws, row, display, cat_key, "expense",
                        fy_labels, cy_labels, cy_start_col, agg)

    _write_total_row(
        ws, OPEX_TOTAL_ROW, "Total Operating Expenses",
        [r for _, (r, _) in OPEX_ROWS.items()],
        fy_labels, cy_labels, cy_start_col, "expense_total",
    )

    # ── NOI (computed) ────────────────────────────────────────────────
    _write_computed_row(ws, NOI_ROW, "NOI (Net Operating Income)",
                        f"={{cell({INCOME_TOTAL_ROW})}}-{{cell({OPEX_TOTAL_ROW})}}",
                        fy_labels, cy_labels, cy_start_col, "noi")
    _write_computed_row(ws, NOI_MARGIN_ROW, "NOI Margin %",
                        f"=IFERROR({{cell({NOI_ROW})}}/{{cell({INCOME_TOTAL_ROW})}},\"-\")",
                        fy_labels, cy_labels, cy_start_col, "noi_margin")

    # ── Utilities ─────────────────────────────────────────────────────
    for cat_key, (row, display) in UTILITY_ROWS.items():
        _write_data_row(ws, row, display, cat_key, "expense",
                        fy_labels, cy_labels, cy_start_col, agg)

    _write_total_row(
        ws, UTILITY_TOTAL_ROW, "Total Utilities",
        [r for _, (r, _) in UTILITY_ROWS.items()],
        fy_labels, cy_labels, cy_start_col, "expense_total",
    )

    # ── Financing ─────────────────────────────────────────────────────
    for cat_key, (row, display) in FINANCING_ROWS.items():
        _write_data_row(ws, row, display, cat_key, "expense",
                        fy_labels, cy_labels, cy_start_col, agg)

    _write_total_row(
        ws, FINANCING_TOTAL_ROW, "Total Financing Cost",
        [r for _, (r, _) in FINANCING_ROWS.items()],
        fy_labels, cy_labels, cy_start_col, "expense_total",
    )

    # ── Capital Allowances ────────────────────────────────────────────
    for cat_key, (row, display) in CAPITAL_ROWS.items():
        _write_data_row(ws, row, display, cat_key, "expense",
                        fy_labels, cy_labels, cy_start_col, agg)

    _write_total_row(
        ws, CAPITAL_TOTAL_ROW, "Total Capital Allowances",
        [r for _, (r, _) in CAPITAL_ROWS.items()],
        fy_labels, cy_labels, cy_start_col, "expense_total",
    )

    # ── Net Profit ────────────────────────────────────────────────────
    _write_net_profit_row(ws, fy_labels, cy_labels, cy_start_col)

    # ── Cash Flow ─────────────────────────────────────────────────────
    for cat_key, (row, display) in CASHFLOW_ROWS.items():
        _write_data_row(ws, row, display, cat_key, "cashflow",
                        fy_labels, cy_labels, cy_start_col, agg)

    _write_total_row(
        ws, NET_CASHFLOW_ROW, "Net Cash Flow",
        [r for _, (r, _) in CASHFLOW_ROWS.items()],
        fy_labels, cy_labels, cy_start_col, "cashflow_total",
    )

    # Principal Repaid (standalone data row, not summed into cash flow)
    _write_data_row(ws, PRINCIPAL_ROW, "Principal Repaid", "principal_repaid",
                    "cashflow", fy_labels, cy_labels, cy_start_col, agg,
                    # Principal is informational, not in SUM
                    )

    # ── KPI Section ───────────────────────────────────────────────────
    _build_kpi_section(ws, fy_labels, cy_labels, cy_start_col)

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    for fi in range(num_fys):
        ft_col = _fy_total_col(fi)
        ws.column_dimensions[get_column_letter(ft_col)].width = 14
        for mi in range(12):
            mc = _month_col(fi, mi)
            ws.column_dimensions[get_column_letter(mc)].width = 11
    for ci in range(len(cy_labels)):
        cc = cy_start_col + ci
        ws.column_dimensions[get_column_letter(cc)].width = 14

    # ── Freeze panes at B5 ────────────────────────────────────────────
    ws.freeze_panes = "B5"

    # ── Print setup ───────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    last_col_letter = get_column_letter(total_cols)
    ws.print_area = f"A1:{last_col_letter}54"
    ws.print_title_rows = "1:4"
    ws.print_title_cols = "A:A"
    ws.oddHeader.left.text = prop.display_name
    ws.oddHeader.right.text = "&D"
    ws.oddFooter.center.text = "Page &P of &N"


# ── Row Writers ───────────────────────────────────────────────────────

def _write_section_header(ws, row: int, text: str, total_cols: int):
    """Write a section header row (INCOME, OPEX, etc.)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = _fill("section_bg")
    cell.font = Font(name="Arial", size=9, bold=True, color=_rgb(COLORS["section_font"]))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def _write_data_row(
    ws, row: int, display_name: str, cat_key: str,
    section_type: str,  # "income" | "expense" | "cashflow"
    fy_labels: list[str], cy_labels: list[str],
    cy_start_col: int, agg: dict,
):
    """Write a data row with values from the aggregation dict."""
    num_fys = len(fy_labels)

    # Style maps
    style_map = {
        "income":   ("income_bg",   "income_font"),
        "expense":  ("expense_bg",  "expense_font"),
        "cashflow": ("cashflow_bg", "cashflow_font"),
    }
    bg_key, font_key = style_map[section_type]

    # Col A: label
    label_cell = ws.cell(row=row, column=1, value=display_name)
    label_cell.fill = _fill(bg_key)
    label_cell.font = _font(font_key, size=9)
    label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 16

    # Month data cells
    for fi, fy in enumerate(fy_labels):
        for mi in range(12):
            mc = _month_col(fi, mi)
            month_key = _month_key_from_fy(fy, mi)
            raw_val = agg.get((cat_key, month_key), 0)

            # In aggregate, expenses are negative. For display,
            # income stays positive; expenses become abs (formulas subtract).
            # Cash flow: cash_received positive, "Less:" items negative in cell.
            if section_type == "expense":
                val = abs(raw_val) if raw_val != 0 else 0
            elif section_type == "cashflow":
                # cash_received was negated by aggregate (it has type cash_flow)
                # but should be positive in sheet; "Less:" items should be negative.
                if cat_key == "cash_received":
                    val = abs(raw_val) if raw_val != 0 else 0
                else:
                    # utilities_paid, mortgage_repayment, capex → negative display
                    val = -abs(raw_val) if raw_val != 0 else 0
            else:
                val = raw_val if raw_val != 0 else 0

            data_cell = ws.cell(row=row, column=mc, value=val if val != 0 else None)
            data_cell.fill = _fill(bg_key)
            data_cell.font = _font(font_key, size=9)
            data_cell.number_format = NUM_FMT_CURRENCY
            data_cell.alignment = Alignment(horizontal="right")

        # FY Total column: SUM of the 12 month columns
        ft_col = _fy_total_col(fi)
        first_mc = _month_col(fi, 0)
        last_mc = _month_col(fi, 11)
        fl = get_column_letter(first_mc)
        ll = get_column_letter(last_mc)
        formula = f"=SUM({fl}{row}:{ll}{row})"

        ft_cell = ws.cell(row=row, column=ft_col, value=formula)
        ft_cell.fill = _fill("fy_total_bg")
        ft_cell.font = Font(name="Arial", size=9, bold=True, color="FF000000")
        ft_cell.number_format = NUM_FMT_CURRENCY
        ft_cell.alignment = Alignment(horizontal="right")

    # CY columns: aggregate from FY blocks that overlap each CY
    for ci, cy in enumerate(cy_labels):
        cc = cy_start_col + ci
        cy_int = int(cy)
        # CY = Jan–Dec of that year.
        # Jan–Jun belongs to one FY, Jul–Dec to another.
        # Build a SUM formula referencing the specific month cells.
        refs = []
        for fi, fy in enumerate(fy_labels):
            start_yr, end_yr = get_fy_year_range(fy)
            for mi in range(12):
                mk = _month_key_from_fy(fy, mi)
                mk_year = int(mk.split("-")[0])
                if mk_year == cy_int:
                    mc = _month_col(fi, mi)
                    refs.append(f"{get_column_letter(mc)}{row}")
        if refs:
            cy_formula = f"=SUM({','.join(refs)})"
        else:
            cy_formula = 0
        cy_cell = ws.cell(row=row, column=cc, value=cy_formula)
        cy_cell.fill = _fill("fy_total_bg")
        cy_cell.font = Font(name="Arial", size=9, bold=True, color="FF000000")
        cy_cell.number_format = NUM_FMT_CURRENCY
        cy_cell.alignment = Alignment(horizontal="right")


def _write_total_row(
    ws, row: int, display_name: str,
    data_rows: list[int],
    fy_labels: list[str], cy_labels: list[str],
    cy_start_col: int,
    style_key: str,  # "income_total" | "expense_total" | "cashflow_total"
):
    """Write a total/summary row with SUM formulas over the data rows."""
    style_map = {
        "income_total":   ("income_total_bg", "income_font"),
        "expense_total":  ("expense_total_bg", "expense_font"),
        "cashflow_total": ("cashflow_total_bg", "cashflow_font"),
    }
    bg_key, font_key = style_map[style_key]
    num_fys = len(fy_labels)

    # Col A
    label_cell = ws.cell(row=row, column=1, value=display_name)
    label_cell.fill = _fill(bg_key)
    label_cell.font = _font(font_key, size=9, bold=True)
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    label_cell.border = BORDER_TOP_THIN_BOTTOM_DOUBLE
    ws.row_dimensions[row].height = 18

    # Build SUM references for each column
    all_cols = []
    for fi in range(num_fys):
        all_cols.append(_fy_total_col(fi))
        for mi in range(12):
            all_cols.append(_month_col(fi, mi))
    for ci in range(len(cy_labels)):
        all_cols.append(cy_start_col + ci)

    for col in all_cols:
        cl = get_column_letter(col)
        parts = [f"{cl}{r}" for r in data_rows]
        formula = f"=SUM({','.join(parts)})"
        cell = ws.cell(row=row, column=col, value=formula)
        cell.fill = _fill(bg_key)
        cell.font = _font(font_key, size=9, bold=True)
        cell.number_format = NUM_FMT_CURRENCY
        cell.alignment = Alignment(horizontal="right")
        cell.border = BORDER_TOP_THIN_BOTTOM_DOUBLE


def _write_computed_row(
    ws, row: int, display_name: str,
    formula_template: str,
    fy_labels: list[str], cy_labels: list[str],
    cy_start_col: int,
    row_type: str,  # "noi" | "noi_margin"
):
    """Write a computed row (NOI, NOI Margin) with per-column formulas."""
    num_fys = len(fy_labels)

    is_margin = row_type == "noi_margin"
    num_fmt = NUM_FMT_PERCENT if is_margin else NUM_FMT_CURRENCY

    # Col A
    label_cell = ws.cell(row=row, column=1, value=display_name)
    label_cell.fill = _fill("noi_bg")
    label_cell.font = _font("noi_font", size=9, bold=True, italic=is_margin)
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    if not is_margin:
        label_cell.border = BORDER_TOP_BOTTOM_MEDIUM
    ws.row_dimensions[row].height = 18 if is_margin else 20

    all_cols = []
    for fi in range(num_fys):
        all_cols.append(_fy_total_col(fi))
        for mi in range(12):
            all_cols.append(_month_col(fi, mi))
    for ci in range(len(cy_labels)):
        all_cols.append(cy_start_col + ci)

    for col in all_cols:
        cl = get_column_letter(col)
        # Replace {cell(ROW)} with actual cell references
        formula = formula_template

        def _replace_cell_ref(m):
            ref_row = m.group(1)
            return f"{cl}{ref_row}"
        formula = re.sub(r'\{cell\((\d+)\)\}', _replace_cell_ref, formula)

        cell = ws.cell(row=row, column=col, value=formula)
        cell.fill = _fill("noi_bg")
        cell.font = _font("noi_font", size=9, bold=True, italic=is_margin)
        cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal="right")
        if not is_margin:
            cell.border = BORDER_TOP_BOTTOM_MEDIUM


def _write_net_profit_row(ws, fy_labels, cy_labels, cy_start_col):
    """NET PROFIT / (LOSS) = NOI - Utilities - Financing - Capital Allowances."""
    row = NET_PROFIT_ROW
    num_fys = len(fy_labels)

    label_cell = ws.cell(row=row, column=1, value="NET PROFIT / (LOSS)")
    label_cell.fill = _fill("noi_bg")
    label_cell.font = Font(name="Arial", size=10, bold=True, color=_rgb(COLORS["noi_font"]))
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    label_cell.border = BORDER_TOP_BOTTOM_THICK
    ws.row_dimensions[row].height = 20

    all_cols = []
    for fi in range(num_fys):
        all_cols.append(_fy_total_col(fi))
        for mi in range(12):
            all_cols.append(_month_col(fi, mi))
    for ci in range(len(cy_labels)):
        all_cols.append(cy_start_col + ci)

    for col in all_cols:
        cl = get_column_letter(col)
        formula = (
            f"={cl}{NOI_ROW}-{cl}{UTILITY_TOTAL_ROW}"
            f"-{cl}{FINANCING_TOTAL_ROW}-{cl}{CAPITAL_TOTAL_ROW}"
        )
        cell = ws.cell(row=row, column=col, value=formula)
        cell.fill = _fill("noi_bg")
        cell.font = Font(name="Arial", size=10, bold=True, color=_rgb(COLORS["noi_font"]))
        cell.number_format = NUM_FMT_CURRENCY
        cell.alignment = Alignment(horizontal="right")
        cell.border = BORDER_TOP_BOTTOM_THICK


# ── KPI Section ───────────────────────────────────────────────────────

def _build_kpi_section(ws, fy_labels, cy_labels, cy_start_col):
    """Build the KPI summary section below the P&L data."""
    num_fys = len(fy_labels)
    total_cols = cy_start_col + len(cy_labels) - 1 if cy_labels else 1 + num_fys * FY_BLOCK_WIDTH

    # Row 57: KPI Header
    kpi_hdr_row = KPI_HEADER_ROW
    ws.merge_cells(start_row=kpi_hdr_row, start_column=1,
                   end_row=kpi_hdr_row, end_column=total_cols)
    kpi_cell = ws.cell(row=kpi_hdr_row, column=1, value="KEY PERFORMANCE INDICATORS")
    kpi_cell.fill = _fill("header_bg")
    kpi_cell.font = Font(name="Arial", size=10, bold=True, color=_rgb(COLORS["header_font"]))
    kpi_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[kpi_hdr_row].height = 22

    # Row 58: KPI column headers (FY labels)
    kpi_col_row = KPI_START_ROW
    ws.cell(row=kpi_col_row, column=1, value="Metric").fill = _fill("section_bg")
    ws.cell(row=kpi_col_row, column=1).font = Font(
        name="Arial", size=8, bold=True, color=_rgb(COLORS["section_font"]))
    ws.cell(row=kpi_col_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=kpi_col_row, column=1).border = Border(bottom=THIN_SIDE)

    for fi, fy in enumerate(fy_labels):
        col = 2 + fi  # KPI uses simple columns: B, C, D, ...
        cell = ws.cell(row=kpi_col_row, column=col, value=fy)
        cell.fill = _fill("section_bg")
        cell.font = Font(name="Arial", size=8, bold=True, color=_rgb(COLORS["section_font"]))
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN_SIDE)

    # KPI data rows
    kpi_metrics = [
        ("Gross Rental Income", INCOME_TOTAL_ROW, NUM_FMT_CURRENCY),
        ("NOI", NOI_ROW, NUM_FMT_CURRENCY),
        ("NOI Margin %", NOI_MARGIN_ROW, NUM_FMT_PERCENT),
        ("Financing Cost", FINANCING_TOTAL_ROW, NUM_FMT_CURRENCY),
        ("Net Profit / (Loss)", NET_PROFIT_ROW, NUM_FMT_CURRENCY),
        ("Net Cash Flow", NET_CASHFLOW_ROW, NUM_FMT_CURRENCY),
        ("DSCR", None, NUM_FMT_RATIO),  # Special formula
    ]

    for ki, (metric_name, src_row, fmt) in enumerate(kpi_metrics):
        data_row = KPI_START_ROW + 1 + ki
        is_alt = ki % 2 == 0
        bg = "kpi_bg" if is_alt else "kpi_alt_bg"

        label_cell = ws.cell(row=data_row, column=1, value=metric_name)
        label_cell.fill = _fill(bg)
        label_cell.font = _font("noi_font", size=9)
        label_cell.alignment = Alignment(horizontal="left", indent=1)

        for fi in range(num_fys):
            col = 2 + fi
            ft_col_letter = get_column_letter(_fy_total_col(fi))
            if src_row is not None:
                formula = f"={ft_col_letter}{src_row}"
            else:
                # DSCR = NOI / (Financing + Principal)
                formula = (
                    f"=IFERROR({ft_col_letter}{NOI_ROW}/"
                    f"({ft_col_letter}{FINANCING_TOTAL_ROW}"
                    f"+{ft_col_letter}{PRINCIPAL_ROW}),\"-\")"
                )

            cell = ws.cell(row=data_row, column=col, value=formula)
            cell.fill = _fill(bg)
            cell.font = _font("noi_font", size=9)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")


# ══════════════════════════════════════════════════════════════════════
#  SUMMARY SHEET BUILDER
# ══════════════════════════════════════════════════════════════════════

def _build_summary_sheet(ws, props, fy_labels, cy_labels):
    """Build the Summary (portfolio dashboard) sheet."""
    num_props = len(props)
    num_fys = len(fy_labels)

    # ── Title (Row 1) ─────────────────────────────────────────────────
    max_col = max(11, 1 + num_fys + len(cy_labels))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title = ws.cell(row=1, column=1, value="Propfolio — Portfolio Summary")
    title.fill = _fill("title_bg")
    title.font = Font(name="Arial", size=12, bold=True, color=_rgb(COLORS["title_font"]))
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ══ Table B: Property Asset Information (Rows 3–) ═════════════════
    table_b_header_row = 3
    table_b_headers = [
        "Property", "Address", "Purchase Price ($)", "Purchase Date",
        "Current Value ($)", "Mortgage Balance ($)", "Equity ($)",
        "LVR (%)", "Gross Yield (%)", "Net Yield (%)", "DSCR",
    ]

    for ci, hdr in enumerate(table_b_headers):
        col = ci + 1
        cell = ws.cell(row=table_b_header_row, column=col, value=hdr)
        cell.fill = _fill("header_bg")
        cell.font = Font(name="Arial", size=9, bold=True, color=_rgb(COLORS["header_font"]))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_SIDE)

    for pi, prop in enumerate(props):
        data_row = table_b_header_row + 1 + pi
        sheet_name = f"IP#{prop.id.replace('IP', '')}"

        # Columns A–F: static property data
        ws.cell(row=data_row, column=1, value=prop.display_name)
        ws.cell(row=data_row, column=2, value=prop.address)
        pp_cell = ws.cell(row=data_row, column=3, value=prop.purchase_price or 0)
        pp_cell.number_format = NUM_FMT_CURRENCY
        pd_cell = ws.cell(row=data_row, column=4, value=prop.purchase_date)
        pd_cell.number_format = "DD/MM/YYYY"
        cv_cell = ws.cell(row=data_row, column=5, value=prop.current_value or 0)
        cv_cell.number_format = NUM_FMT_CURRENCY
        mb_cell = ws.cell(row=data_row, column=6, value=prop.mortgage_balance or 0)
        mb_cell.number_format = NUM_FMT_CURRENCY

        r = data_row
        # Equity = Current Value - Mortgage
        ws.cell(row=r, column=7, value=f"=E{r}-F{r}").number_format = NUM_FMT_CURRENCY
        # LVR
        ws.cell(row=r, column=8, value=f'=IFERROR(F{r}/E{r},"-")').number_format = NUM_FMT_PERCENT
        # Gross Yield: FY Rent / Current Value
        fy_total_letter = get_column_letter(_fy_total_col(0))
        ws.cell(row=r, column=9,
                value=f"=IFERROR('{sheet_name}'!{fy_total_letter}{INCOME_TOTAL_ROW}/E{r},\"-\")"
                ).number_format = NUM_FMT_PERCENT
        # Net Yield: FY NOI / Current Value
        ws.cell(row=r, column=10,
                value=f"=IFERROR('{sheet_name}'!{fy_total_letter}{NOI_ROW}/E{r},\"-\")"
                ).number_format = NUM_FMT_PERCENT
        # DSCR
        ws.cell(row=r, column=11,
                value=(
                    f"=IFERROR('{sheet_name}'!{fy_total_letter}{NOI_ROW}/"
                    f"('{sheet_name}'!{fy_total_letter}{FINANCING_TOTAL_ROW}"
                    f"+'{sheet_name}'!{fy_total_letter}{PRINCIPAL_ROW}),\"-\")"
                )).number_format = NUM_FMT_RATIO

        # Style alternating rows
        bg = "kpi_bg" if pi % 2 == 0 else "kpi_alt_bg"
        for c in range(1, 12):
            ws.cell(row=data_row, column=c).fill = _fill(bg)

    # ══ Table A: Portfolio Performance (starts after Table B) ═════════
    perf_start_row = table_b_header_row + num_props + 3  # 2 blank rows gap

    # Performance section header
    ws.merge_cells(start_row=perf_start_row, start_column=1,
                   end_row=perf_start_row, end_column=max_col)
    perf_title = ws.cell(row=perf_start_row, column=1,
                         value="PORTFOLIO PERFORMANCE BY FINANCIAL YEAR")
    perf_title.fill = _fill("title_bg")
    perf_title.font = Font(name="Arial", size=10, bold=True, color=_rgb(COLORS["title_font"]))
    perf_title.alignment = Alignment(horizontal="center")

    # Column headers for performance: A=Metric, then one col per FY, then CY cols
    col_hdr_row = perf_start_row + 1

    ws.cell(row=col_hdr_row, column=1, value="Metric").fill = _fill("header_bg")
    ws.cell(row=col_hdr_row, column=1).font = Font(
        name="Arial", size=9, bold=True, color=_rgb(COLORS["header_font"]))

    for fi, fy in enumerate(fy_labels):
        col = 2 + fi
        cell = ws.cell(row=col_hdr_row, column=col, value=fy)
        cell.fill = _fill("header_bg")
        cell.font = Font(name="Arial", size=9, bold=True, color=_rgb(COLORS["header_font"]))
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN_SIDE)

    # Performance metrics: For each metric, one row per property + Portfolio Total
    perf_metrics = [
        ("Gross Rental Income", INCOME_TOTAL_ROW, NUM_FMT_CURRENCY),
        ("NOI", NOI_ROW, NUM_FMT_CURRENCY),
        ("NOI Margin %", NOI_MARGIN_ROW, NUM_FMT_PERCENT),
        ("Financing Cost", FINANCING_TOTAL_ROW, NUM_FMT_CURRENCY),
        ("Net Profit / (Loss)", NET_PROFIT_ROW, NUM_FMT_CURRENCY),
        ("Net Cash Flow", NET_CASHFLOW_ROW, NUM_FMT_CURRENCY),
    ]

    current_row = col_hdr_row + 1
    for metric_name, src_row, fmt in perf_metrics:
        # Metric sub-header
        mh_cell = ws.cell(row=current_row, column=1, value=metric_name)
        mh_cell.fill = _fill("section_bg")
        mh_cell.font = Font(name="Arial", size=9, bold=True, color=_rgb(COLORS["section_font"]))
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=1 + num_fys)
        current_row += 1

        prop_data_rows = []
        for pi, prop in enumerate(props):
            sheet_name = f"IP#{prop.id.replace('IP', '')}"
            ws.cell(row=current_row, column=1, value=f"  {prop.display_name}")

            for fi in range(num_fys):
                col = 2 + fi
                ft_letter = get_column_letter(_fy_total_col(fi))
                formula = f"='{sheet_name}'!{ft_letter}{src_row}"
                cell = ws.cell(row=current_row, column=col, value=formula)
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")

                bg = "kpi_bg" if pi % 2 == 0 else "kpi_alt_bg"
                cell.fill = _fill(bg)

            ws.cell(row=current_row, column=1).fill = _fill(
                "kpi_bg" if pi % 2 == 0 else "kpi_alt_bg")

            prop_data_rows.append(current_row)
            current_row += 1

        # Portfolio Total row
        ws.cell(row=current_row, column=1, value="Portfolio Total").font = Font(
            name="Arial", size=9, bold=True)
        ws.cell(row=current_row, column=1).fill = _fill("income_total_bg")

        for fi in range(num_fys):
            col = 2 + fi
            cl = get_column_letter(col)
            if fmt == NUM_FMT_PERCENT:
                # Average for percentages
                refs = ",".join(f"{cl}{r}" for r in prop_data_rows)
                formula = f'=IFERROR(AVERAGE({refs}),"-")'
            else:
                refs = ",".join(f"{cl}{r}" for r in prop_data_rows)
                formula = f"=SUM({refs})"
            cell = ws.cell(row=current_row, column=col, value=formula)
            cell.number_format = fmt
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.fill = _fill("income_total_bg")
            cell.alignment = Alignment(horizontal="right")
            cell.border = BORDER_TOP_THIN_BOTTOM_DOUBLE

        current_row += 1  # blank row between metrics

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    for ci in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    ws.freeze_panes = "B2"
